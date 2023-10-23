import pandas as pd
import os
import shutil
import openpyxl

dataDir = os.getcwd()
data = dataDir + "\\train_val_test set.xlsx"
leftImage = dataDir + "\\left_imgs_cropped_resized"
rightImage = dataDir + "\\right_imgs_cropped"

leftTrain = dataDir + "\\leftTrain"
rightTrain = dataDir + "\\rightTrain"
leftVal = dataDir + "\\leftVal"
rightVal = dataDir + "\\rightVal"


# the total number of photo is 1022
#      folder                            total        left      right
# the number of images in folder 2014:    402          201       201
#                      in folder GE  :    435          223       212
#                      in folder phil:    185          97        88
# the total number                   :   1022          521       501

# def splitLR():
#     cwd = os.getcwd()
#     cwd += "\\all"
#     # left = []
#     all_photo = os.listdir(cwd)
#
#     # split names into two lists
#     right = [r for r in all_photo if "r" in r or "R" in r]
#     left = [l for l in all_photo if l not in right]
#     return left, len(right)


# the function move into designated folder and return all the images names in one list
def intoDir(des):
    cwd = dataDir + "\\" + des
    name = os.listdir(cwd)
    return name


# the function separates the original images from three folders into left and right
def splitLR():
    # the destinations of left and right images
    leftDes = dataDir + '\\left'
    os.makedirs(leftDes) # make two directories to store images
    rightDes = dataDir + '\\right'
    os.makedirs(rightDes)

    # run through three directories to split images in them one by one
    folder = ['deidentifiedImages_2014', 'deidentifiedImages_GE', 'deidentifiedImages_phillips']
    # folder = ['deidentifiedImages_GE']

    for f in folder:
        name = intoDir(f)  # the list contains names of the images in the directory
        countL = 0
        countR = 0
        total = len(name)

        # if f == 'deidentifiedImages_2014':
        if '2014' in f:
            right = [r for r in name if "r" in r or "R" in r]
            left = [l for l in name if l not in right]
            countL = len(left)
            countR = len(right)

            # for l in left:
            #     shutil.move(dataDir + '\\' + f + '\\' + l, leftDes + '\\' + l[:3] + '.png')
            # for r in right:
            #     shutil.move(dataDir + '\\' + f + '\\' + r, rightDes + '\\' + r[:3] + '.png')
        # there is a problem with patient 2011296 since the image has no clear class
        # i put it into the left class
        elif f == 'deidentifiedImages_GE' or f == 'deidentifiedImages_phillips':
            left = [l for l in name if 'lt' in l or 'LT' in l]
            right2 = [r for r in name if 'rt' in r or 'RT' in r]
            right = [r for r in name if r not in left]
            # print(len(right2))
            countL = len(left)
            countR = len(right)
        for l in left:
            shutil.move(dataDir + '\\' + f + '\\' + l, leftDes + '\\' + l)
        for r in right:
            shutil.move(dataDir + '\\' + f + '\\' + r, rightDes + '\\' + r)
        print("{} has {} images, are seperated into left and right with {} in left and {} in right".format(f, total, countL, countR))


# Death/ACS/Stroke/TIA/CEA/CABG/PCI/CHF   Death  MI  ACS  Unstable Angina  Stroke  TIA  CABG  PCI  HF
def get_label():  # return the 2d array of the image and its class
    cwd = os.getcwd()  # cwd is the directory this file in
    data_root = os.path.join(cwd, "CD.xlsx")
    # print(data_root)
    workbook = openpyxl.load_workbook("CD.xlsx")

    # Define variable to read the active sheet:
    worksheet = workbook.active
    label = []
    conciseL = []
    for i in range(1, worksheet.max_row):
        flag = 0
        temp = []
        for col in worksheet.iter_cols(1, 3):
            if flag != 1:
                temp.append(col[i].value)
            if flag == 2:
                conciseL.append(col[i].value)
            flag += 1
        label.append(temp)
        # print(col[i].value, end="\t\t")
    # print(len(label[:-8]))
    return label[:-8], conciseL[:-8]


# the function split the images into train, val, test folders
def split(origin: str, side: str, type: str):
    # origin is the folder of origin images, side is the side of the bulb, type = [train, val, test]
    des = dataDir + "\\{}{}".format(side, type)
    print(des)
    if os.path.exists(des):
        shutil.rmtree(des)
    os.makedirs(des)

    name = get_list('{} Data - {} Bulb'.format(type, side), 'label_BU')[1]
    # print(name)
    for n in name:
        try:
            shutil.move(origin + "\\{}".format(n), des + "\\{}".format(n))
        except:
            print(n)


def splitCls(origin, side, t):
    des = dataDir + "\\{}{}{}".format(side, t, '0')
    des1 = dataDir + "\\{}{}{}".format(side, t, '1')
    if os.path.exists(des):
        shutil.rmtree(des)
    if os.path.exists(des1):
        shutil.rmtree(des1)
    os.mkdir(des)
    os.mkdir(des1)
    # print(des)
    # print(des1)

    name = os.listdir(origin)
    label = get_list('{} Data - {} Bulb'.format(t, side), 'label_BU')[3]
    # print(type(label))
    # print(label)
    c0 = 0
    c1 = 0
    for n in name:
        print(label[n])
        if label[n] == 0:
            shutil.move(origin + "\\{}".format(n), des + "\\{}".format(n))
            c0 += 1
        elif label[n] == 1:
            shutil.move(origin + "\\{}".format(n), des1 + "\\{}".format(n))
            c1 += 1
    print(c0, c1)


# the function returns a 2D list, each element contains the name of the image and the class
def get_list(sheet: str, col: str):
    carotid_data = pd.read_excel(data, sheet_name=sheet, usecols="A,B", engine='openpyxl')
    # print(pd.ExcelFile(data).sheet_names)
    # ['Train Data - Right Bulb', 'Val Data - Right Bulb', 'Test Data - Right Bulb',
    # 'Train Data - Left Bulb', 'Val Data - Left Bulb', 'Test Data - Left Bulb']

    # convert the data type from frame work to list
    n = carotid_data[["filename"]].values.tolist()
    name = []
    for e in n:
        name.append(e[0])

    c = carotid_data[[col]].values.tolist()
    cla = []
    for e in c:
        cla.append(e[0])  # the data type is int

    # carotid_data = carotid_data.to_dict()
    carotid_data = carotid_data.values.tolist()

    # print(name, cla)
    dic = {}
    for n, c in zip(name, cla):
        dic[n] = c

    return carotid_data, name, cla, dic


# the function should read the images from folder and label form excel
def match():
    train = []
    trainLabel = []
    val = []
    valLabel = []

    trainDic = get_list('Train Data - Right Bulb', 'label_BU')[3]
    valDic = get_list('Val Data - Right Bulb', 'label_BU')[3]

    tr = os.listdir(rightTrain) # train
    for t in tr:
        train.append(rightTrain + "\\{}".format(t))
        trainLabel.append(trainDic[t])

    va = os.listdir(rightVal) # val
    for v in va:
        val.append(rightVal + "\\{}".format(v))
        valLabel.append(valDic[v])

    return train, trainLabel, val, valLabel


if __name__ == "__main__":
    # get_list('Train Data - Left Bulb', 'label_BU')
    # split(leftImage, "Left", "Train") # suppose to be 375, only has 374
    # split(rightImage, "Right", "Train")
    # splitCls(rightVal, "Right", "Val")
    # print(data)
    print(match())
